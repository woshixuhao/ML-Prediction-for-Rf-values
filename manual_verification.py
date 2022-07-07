import matplotlib.pyplot as plt
from PIL import Image
import os
import openpyxl
import pandas as pd

def manual_vertification(path):
    '''
    :param path: The dir tha saves pictures needed to be vertified
    :return:
    '''
    # Get all files in this directory and store them in the list
    filelist = os.listdir(path)
    filelist = pd.DataFrame(filelist)
    filelist.columns = ["A"]
    filelist['B'] = filelist.loc[:, 'A']
    filelist['C'] = filelist.loc[:, 'A']
    for i in range(len(filelist)):
        filelist.loc[i, 'B'] = int(filelist.loc[i, 'A'].split('_', )[1])
        filelist.loc[i, 'C'] = int(filelist.loc[i, 'A'].split('.', )[0].split('_', )[-1])
    filelist.sort_values(by=['B', 'C'], axis=0, inplace=True)
    print(filelist)

    for epoch in range(len(filelist)):
        '''
        获得TLC板的信息，包括化合物编号以及展开剂
        '''
        TLC_name = filelist.loc[:, 'A'][epoch]
        compound_id = []
        for i in range(4):
            compound_id.append(int(TLC_name.split('_', )[1]) + i)

        eluent = TLC_name.split('.', )[0].split('_', )[-1]

        plt.figure(figsize=(10, 16))
        img = Image.open(path + os.sep + TLC_name)
        plt.imshow(img)  # 打开图片
        '''
    Click the points of the four compounds in turn, 
    as well as the lower edge of the TLC plate and the front edge of the solvent, 
    and the Rf value is automatically calculated
        '''
        distance = 110  # Pixel distance from the origin to the bottom edge of the board

        compound_spot = plt.ginput(4, timeout=90)
        start_end = plt.ginput(2, timeout=90)
        plt.close()
        # 打开文档，写入数据
        wb = openpyxl.load_workbook(r'C:\Users\mofan\Documents\My Document\07 Papers\9 徐浩\data\TLC_data.xlsx')
        ws = wb.active

        for i in range(4):
            Rf = (compound_spot[i][1] - start_end[0][1] + distance) / (start_end[1][1] - start_end[0][1] + distance)
            print(Rf)
            compound_cell = 'A' + str(compound_id[i]+1)
            ws[compound_cell] = str(compound_id[i])
            Rf_cell = chr(ord(str(eluent)) + 17) + str(compound_id[i]+1)
            ws[Rf_cell] = Rf

        wb.save(r'C:\Users\mofan\Documents\My Document\07 Papers\9 徐浩\data\TLC_data.xlsx')
