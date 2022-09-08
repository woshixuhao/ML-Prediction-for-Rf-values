''''
ver2021.12.2 正式版点板，一次3个点板罐子
拿板子间隔12分33秒
更改了命名规则
去除了反面拍照
'''''
from camera import *
from skimage.io import imread, imshow
from skimage.color import rgb2gray, rgb2hsv
import cv2
import numpy as np
import math
import DobotSDK as dobotSDK
from threading import Timer
import matplotlib.pyplot as plt
import time
import openpyxl
import datetime
from camera import *
from skimage.io import imread, imshow
from skimage.color import rgb2gray, rgb2hsv
import cv2
import numpy as np
import math
import DobotSDK as dobotSDK
from threading import Timer
import matplotlib.pyplot as plt
import warnings

warnings.filterwarnings("ignore")
# 将dll读取到内存中并获取对应的CDLL实例
# Load Dll and get the CDLL object
class TimerClass():
    api = 0

    def __init__(self, api):
        self.api = api

    def __del__(self):
        del self.api

    def printExchange(self, inc):
        global timer_running
        if timer_running == True:
            t = Timer(inc, self.printExchange, (inc,))
            t.start()
            exchangeList = dobotSDK.GetExchange(self.api)
            print("*************************************")
            print("ControlMode:", exchangeList[1])
            print("PowerState:", exchangeList[2])
            print("IsCollision:", exchangeList[3])
            print("JogMode:", exchangeList[4])
            print("IsAuto:", exchangeList[5])
            print("ToolCoordinate:", exchangeList[6])
            print("UserCoordinate:", exchangeList[7])
            print("Joint:", exchangeList[8])
            print("Coordinate:", exchangeList[9])
            print("Alarms:", exchangeList[10])
            print("RDN:", exchangeList[11])
            print("Arm:", exchangeList[12])
            print("DO:", exchangeList[13])
            print("DI:", exchangeList[14])
            print("*************************************")
            if len(exchangeList[10]) != 0:
                t.cancel()
                for id, type in exchangeList[10].items():
                    print("GetAlarmsParameter:", dobotSDK.GetAlarmsParameter(api, id, type, True))


from robotcontrol import *

global griper_id
import serial

gripper_id = 1


def rad_to_deg(rad):
    rad = list(rad)
    for i in range(len(rad)):
        rad[i] = rad[i] * 180 / 3.1415926
    deg = tuple(rad)
    return deg


def imageSimilarity(img, img1):
    H1 = cv2.calcHist([img], [1], None, [256], [0, 256])
    H1 = cv2.normalize(H1, H1, 0, 1, cv2.NORM_MINMAX, -1)  # 对图片进行归一化处理

    # 计算图img2的直方图
    H2 = cv2.calcHist([img1], [1], None, [256], [0, 256])
    H2 = cv2.normalize(H2, H2, 0, 1, cv2.NORM_MINMAX, -1)

    # 利用compareHist（）进行比较相似度
    similarity = cv2.compareHist(H1, H2, 0)
    return similarity


def four_to_ola(w, x, y, z):
    r = math.atan2(2 * (w * x + y * z), 1 - 2 * (x * x + y * y))
    r = r / math.pi * 180
    p = math.asin(2 * (w * y - z * x))
    p = p / math.pi * 180
    y = math.atan2(2 * (w * z + x * y), 1 - 2 * (y * y + z * z))
    y = y / math.pi * 180
    return r, p, y


def ola_to_four(r, p, y):
    # 输入为角度，切勿转化为弧度
    sinp = math.sin(math.radians(p / 2))
    siny = math.sin(math.radians(y / 2))
    sinr = math.sin(math.radians(r / 2))

    cosp = math.cos(math.radians(p / 2))
    cosy = math.cos(math.radians(y / 2))
    cosr = math.cos(math.radians(r / 2))

    w = cosr * cosp * cosy + sinr * sinp * siny
    x = sinr * cosp * cosy - cosr * sinp * siny
    y = cosr * sinp * cosy + sinr * cosp * siny
    z = cosr * cosp * siny - sinr * sinp * cosy

    return w, x, y, z


def eye_to_photo(x, y):
    x_hat = 4.82553606e-06 * x + 6.95547182e-05 * y - 7.07863871e-01
    y_hat = 8.02383041e-05 * x - 8.37919035e-06 * y - 1.68856356e-01
    return x_hat, y_hat


# --------------------夹爪函数------------------
global griper_id
gripper_id = 1


# 把数据分成高字节和低字节
def data2bytes(data):
    rdata = [0xff] * 2
    if data == -1:
        rdata[0] = 0xff
        rdata[1] = 0xff
    else:
        rdata[0] = data & 0xff
        rdata[1] = (data >> 8) & (0xff)
    return rdata


# 把十六进制或十进制的数转成bytes
def num2str(num):
    str = hex(num)
    str = str[2:4]
    if (len(str) == 1):
        str = '0' + str
    str = bytes.fromhex(str)
    # print(str)
    return str


# 求校验和
def checknum(data, leng):
    result = 0
    for i in range(2, leng):
        result += data[i]
    result = result & 0xff
    # print(result)
    return result


# 扫描id号
def getid(i):
    global gripper_id

    datanum = 0x05
    b = [0] * (datanum + 5)
    # 包头
    b[0] = 0xEB
    b[1] = 0x90

    # id号
    b[2] = i

    # 数据个数
    b[3] = datanum

    # 操作码
    b[4] = 0x12

    # 数据
    b[5] = data2bytes(1000)[0]
    b[6] = data2bytes(1000)[1]

    b[7] = data2bytes(0)[0]
    b[8] = data2bytes(0)[1]

    # 校验和
    b[9] = checknum(b, datanum + 4)

    # 向串口发送数据
    putdata = b''

    for i in range(1, datanum + 6):
        putdata = putdata + num2str(b[i - 1])
    ser.write(putdata)
    # print('发送的数据：',putdata)

    getdata = ser.read(7)
    return len(getdata)


# 设置开口限位（最大开口度和最小开口度）
def setopenlimit(openmax, openmin):
    global gripper_id
    if openmax < 0 or openmax > 1000:
        print('数据超出正确范围：0-1000')
        return
    if openmin < 0 or openmin > 1000:
        print('数据超出正确范围：0-1000')
        return
    if openmax < openmin:
        print('最大开口度应该大于最小开口度')
        return

    datanum = 0x05
    b = [0] * (datanum + 5)
    # 包头
    b[0] = 0xEB
    b[1] = 0x90

    # id号
    b[2] = gripper_id

    # 数据个数
    b[3] = datanum

    # 操作码
    b[4] = 0x12

    # 数据
    b[5] = data2bytes(openmax)[0]
    b[6] = data2bytes(openmax)[1]

    b[7] = data2bytes(openmin)[0]
    b[8] = data2bytes(openmin)[1]

    # 校验和
    b[9] = checknum(b, datanum + 4)

    # 向串口发送数据
    putdata = b''

    for i in range(1, datanum + 6):
        putdata = putdata + num2str(b[i - 1])
    ser.write(putdata)
    # print('发送的数据：',putdata)

    print('发送的数据：')
    for i in range(1, datanum + 6):
        print(hex(putdata[i - 1]))

    getdata = ser.read(7)
    print('返回的数据：')
    for i in range(1, 8):
        print(hex(getdata[i - 1]))


# 设置ID
def setid(idnew):
    global gripper_id
    if idnew < 1 or idnew > 254:
        print('数据超出正确范围：1-254')
        return

    datanum = 0x02
    b = [0] * (datanum + 5)
    # 包头
    b[0] = 0xEB
    b[1] = 0x90
    # id号
    b[2] = gripper_id

    # 数据个数
    b[3] = datanum

    # 操作码
    b[4] = 0x04

    # 数据
    b[5] = idnew

    gripper_id = idnew

    # 校验和
    b[6] = checknum(b, datanum + 4)

    # 向串口发送数据
    putdata = b''

    for i in range(1, datanum + 6):
        putdata = putdata + num2str(b[i - 1])
    ser.write(putdata)
    # print('发送的数据：',putdata)

    # print('发送的数据：')
    # for i in range(1,datanum+6):
    #     print(hex(putdata[i-1]))
    # getdata= ser.read(7)
    # print('返回的数据：')
    # for i in range(1,8):
    #     print(hex(getdata[i-1]))


# 运动到目标
def movetgt(tgt):
    global gripper_id
    if tgt < 0 or tgt > 1000:
        print('数据超出正确范围：0-1000')
        return

    datanum = 0x03
    b = [0] * (datanum + 5)
    # 包头
    b[0] = 0xEB
    b[1] = 0x90
    # id号
    b[2] = gripper_id

    # 数据个数
    b[3] = datanum

    # 操作码
    b[4] = 0x54

    # 数据
    b[5] = data2bytes(tgt)[0]
    b[6] = data2bytes(tgt)[1]

    # 校验和
    b[7] = checknum(b, datanum + 4)

    # 向串口发送数据
    putdata = b''

    for i in range(1, datanum + 6):
        putdata = putdata + num2str(b[i - 1])
    ser.write(putdata)
    # print('发送的数据：',putdata)

    # print('发送的数据：')
    # for i in range(1, datanum + 6):
    #     print(hex(putdata[i - 1]))
    # getdata = ser.read(7)
    # print('返回的数据：')
    # for i in range(1, 8):
    #     print(hex(getdata[i - 1]))


# 运动张开
def movemax(speed):
    global gripper_id
    if speed < 1 or speed > 1000:
        print('数据超出正确范围：1-1000')
        return

    datanum = 0x03
    b = [0] * (datanum + 5)
    # 包头
    b[0] = 0xEB
    b[1] = 0x90
    # id号
    b[2] = gripper_id

    # 数据个数
    b[3] = datanum

    # 操作码
    b[4] = 0x11

    # 数据
    b[5] = data2bytes(speed)[0]
    b[6] = data2bytes(speed)[1]

    # 校验和
    b[7] = checknum(b, datanum + 4)

    # 向串口发送数据
    putdata = b''

    for i in range(1, datanum + 6):
        putdata = putdata + num2str(b[i - 1])
    ser.write(putdata)
    # print('发送的数据：',putdata)

    # print('发送的数据：')
    # for i in range(1,datanum+6):
    #     print(hex(putdata[i-1]))
    # getdata= ser.read(7)
    # print('返回的数据：')
    # for i in range(1,8):
    #     print(hex(getdata[i-1]))


# 运动闭合
def movemin(speed, power):
    global gripper_id
    if speed < 1 or speed > 1000:
        print('数据超出正确范围：1-1000')
        return
    if power < 50 or speed > 1000:
        print('数据超出正确范围：50-1000')
        return

    datanum = 0x05
    b = [0] * (datanum + 5)
    # 包头
    b[0] = 0xEB
    b[1] = 0x90
    # id号
    b[2] = gripper_id

    # 数据个数
    b[3] = datanum

    # 操作码
    b[4] = 0x10

    # 数据
    b[5] = data2bytes(speed)[0]
    b[6] = data2bytes(speed)[1]
    b[7] = data2bytes(power)[0]
    b[8] = data2bytes(power)[1]
    # 校验和
    b[9] = checknum(b, datanum + 4)

    # 向串口发送数据
    putdata = b''

    for i in range(1, datanum + 6):
        putdata = putdata + num2str(b[i - 1])
    ser.write(putdata)
    # print('发送的数据：',putdata)

    print('发送的数据：')
    for i in range(1, datanum + 6):
        print(hex(putdata[i - 1]))
    getdata = ser.read(7)
    # print('返回的数据：')
    # for i in range(1,8):
    #     print(hex(getdata[i-1]))


# 运动持续闭合
def moveminhold(speed, power):
    global gripper_id
    if speed < 1 or speed > 1000:
        print('数据超出正确范围：1-1000')
        return
    if power < 50 or speed > 1000:
        print('数据超出正确范围：50-1000')
        return

    datanum = 0x05
    b = [0] * (datanum + 5)
    # 包头
    b[0] = 0xEB
    b[1] = 0x90
    # id号
    b[2] = gripper_id

    # 数据个数
    b[3] = datanum

    # 操作码
    b[4] = 0x18

    # 数据
    b[5] = data2bytes(speed)[0]
    b[6] = data2bytes(speed)[1]
    b[7] = data2bytes(power)[0]
    b[8] = data2bytes(power)[1]
    # 校验和
    b[9] = checknum(b, datanum + 4)

    # 向串口发送数据
    putdata = b''

    for i in range(1, datanum + 6):
        putdata = putdata + num2str(b[i - 1])
    ser.write(putdata)
    # print('发送的数据：',putdata)
    #
    # print('发送的数据：')
    # for i in range(1,datanum+6):
    #     print(hex(putdata[i-1]))
    # getdata= ser.read(7)
    # print('返回的数据：')
    # for i in range(1,8):
    #     print(hex(getdata[i-1]))


# 读取开口限位
def getopenlimit():
    global gripper_id

    datanum = 0x01
    b = [0] * (datanum + 5)
    # 包头
    b[0] = 0xEB
    b[1] = 0x90

    # gripper_id号
    b[2] = gripper_id

    # 数据个数
    b[3] = datanum

    # 操作码
    b[4] = 0x13

    # 校验和
    b[5] = checknum(b, datanum + 4)

    # 向串口发送数据
    putdata = b''
    for i in range(1, datanum + 6):
        putdata = putdata + num2str(b[i - 1])
    ser.write(putdata)
    print('发送的数据：')
    for i in range(1, datanum + 6):
        print(hex(putdata[i - 1]))

    getdata = ser.read(10)
    print('返回的数据：')
    for i in range(1, 11):
        print(hex(getdata[i - 1]))

    openlimit = [0] * 2
    for i in range(1, 3):
        if getdata[i * 2 + 3] == 0xff and getdata[i * 2 + 4] == 0xff:
            openlimit[i - 1] = -1
        else:
            openlimit[i - 1] = getdata[i * 2 + 3] + (getdata[i * 2 + 4] << 8)
    return openlimit


# 读取当前开口
def getcopen():
    global gripper_id

    datanum = 0x01
    b = [0] * (datanum + 5)
    # 包头
    b[0] = 0xEB
    b[1] = 0x90

    # gripper_id号
    b[2] = gripper_id

    # 数据个数
    b[3] = datanum

    # 操作码
    b[4] = 0xD9

    # 校验和
    b[5] = checknum(b, datanum + 4)

    # 向串口发送数据
    putdata = b''
    for i in range(1, datanum + 6):
        putdata = putdata + num2str(b[i - 1])
    ser.write(putdata)
    # print('发送的数据：')
    # for i in range(1,datanum+6):
    #     print(hex(putdata[i-1]))

    getdata = ser.read(8)
    # print('返回的数据：')
    # for i in range(1,9):
    #     print(hex(getdata[i-1]))

    copen = [0] * 1
    for i in range(1, 2):
        if getdata[i * 2 + 3] == 0xff and getdata[i * 2 + 4] == 0xff:
            copen[i - 1] = -1
        else:
            copen[i - 1] = getdata[i * 2 + 3] + (getdata[i * 2 + 4] << 8)
    return copen


# 读取当前状态
def getstate():
    global gripper_id

    datanum = 0x01
    b = [0] * (datanum + 5)
    # 包头
    b[0] = 0xEB
    b[1] = 0x90

    # gripper_id号
    b[2] = gripper_id

    # 数据个数
    b[3] = datanum

    # 操作码
    b[4] = 0x41

    # 校验和
    b[5] = checknum(b, datanum + 4)

    # 向串口发送数据
    putdata = b''
    for i in range(1, datanum + 6):
        putdata = putdata + num2str(b[i - 1])
    ser.write(putdata)
    # print('发送的数据：')
    # for i in range(1,datanum+6):
    #     print(hex(putdata[i-1]))

    getdata = ser.read(13)
    # print('返回的数据：')
    # for i in range(1,14):
    #     print(hex(getdata[i-1]))

    if getdata[5] == 1:
        print('max in place')
    elif getdata[5] == 2:
        print('min in place')
    elif getdata[5] == 3:
        print('stop in place')
    elif getdata[5] == 4:
        print('closing')
    elif getdata[5] == 5:
        print('openning')
    elif getdata[5] == 6:
        print('force control in place to stop')
    else:
        print('no def')

    if (getdata[6] & 0x01) == 1:
        print('runing stop fault')

    if (getdata[6] & 0x02) == 2:
        print('overheat fault')

    if (getdata[6] & 0x04) == 4:
        print('Over Current Fault')

    if (getdata[6] & 0x08) == 8:
        print('running fault')

    if (getdata[6] & 0x10) == 16:
        print('communication fault')

    print('temp:', getdata[7])
    print('curopen:', ((getdata[9] << 8) & 0xff00) + getdata[8])
    print('power:', ((getdata[11] << 8) & 0xff00) + getdata[10])


# 急停
def setestop():
    global gripper_id

    datanum = 0x01
    b = [0] * (datanum + 5)
    # 包头
    b[0] = 0xEB
    b[1] = 0x90
    # id号
    b[2] = gripper_id

    # 数据个数
    b[3] = datanum

    # 操作码
    b[4] = 0x16

    # 校验和
    b[5] = checknum(b, datanum + 4)

    # 向串口发送数据
    putdata = b''

    for i in range(1, datanum + 6):
        putdata = putdata + num2str(b[i - 1])
    ser.write(putdata)
    # print('发送的数据：',putdata)

    print('发送的数据：')
    for i in range(1, datanum + 6):
        print(hex(putdata[i - 1]))
    getdata = ser.read(7)
    print('返回的数据：')
    for i in range(1, 8):
        print(hex(getdata[i - 1]))

        # 参数固化


def setparam():
    global gripper_id

    datanum = 0x01
    b = [0] * (datanum + 5)
    # 包头
    b[0] = 0xEB
    b[1] = 0x90
    # id号
    b[2] = gripper_id

    # 数据个数
    b[3] = datanum

    # 操作码
    b[4] = 0x01

    # 校验和
    b[5] = checknum(b, datanum + 4)

    # 向串口发送数据
    putdata = b''

    for i in range(1, datanum + 6):
        putdata = putdata + num2str(b[i - 1])
    ser.write(putdata)
    # print('发送的数据：',putdata)

    print('发送的数据：')
    for i in range(1, datanum + 6):
        print(hex(putdata[i - 1]))
    getdata = ser.read(7)
    print('返回的数据：')
    for i in range(1, 8):
        print(hex(getdata[i - 1]))

        # 清除故障


def setFrsvd():
    global gripper_id

    datanum = 0x01
    b = [0] * (datanum + 5)
    # 包头
    b[0] = 0xEB
    b[1] = 0x90
    # id号
    b[2] = gripper_id

    # 数据个数
    b[3] = datanum

    # 操作码
    b[4] = 0x17

    # 校验和
    b[5] = checknum(b, datanum + 4)

    # 向串口发送数据
    putdata = b''

    for i in range(1, datanum + 6):
        putdata = putdata + num2str(b[i - 1])
    ser.write(putdata)
    # print('发送的数据：',putdata)

    # print('发送的数据：')
    # for i in range(1,datanum+6):
    #     print(hex(putdata[i-1]))
    # getdata= ser.read(7)
    # print('返回的数据：')
    # for i in range(1,8):
    #     print(hex(getdata[i-1]))


def robot_login():
    ret = Auboi5Robot.initialize()
    # 创建一个实例化对象
    robot = Auboi5Robot()

    # 创建一个句柄

    handle = robot.create_context()

    # 打印上下文
    logger.info("robot.rshd={0}".format(handle))

    # 链接服务器
    ip = '192.168.1.2'
    port = 8899
    result = robot.connect(ip, port)

    if result != RobotErrorType.RobotError_SUCC:
        logger.info("connect server{0}:{1} failed.".format(ip, port))
    else:
        collision = 6
        tool_dynamics = {"position": (0, 0, 0), "payload": 0.0, "inertia": (0, 0, 0, 0, 0, 0)}
        ret = robot.robot_startup(collision, tool_dynamics)

        # 关节运动
        # 初始化全局运动属性
        # robot.init_profile()
        # 设置关节最大加速度
        # joint_maxvelc = (2.596177, 2.596177, 2.596177, 3.110177, 3.110177, 3.110177)
        # joint_maxacc = (
        #     17.308779 / 10, 17.308779 / 10, 17.308779 / 10, 17.308779 / 10, 17.308779 / 10, 17.308779 / 10)
        # robot.set_joint_maxacc(joint_maxacc)
        # robot.set_joint_maxvelc(joint_maxvelc)
        # robot.set_arrival_ahead_blend(0.05)
        # joint_radian = (0, 0, 0, 0, 0, 0)
        # ret=robot.move_joint(joint_radian)
        #
        # robot.enable_robot_event()
        # robot.init_profile()
    return robot, ret


def deg_to_rad(deg):
    # deg is 6-shape tuple indicate the location
    deg = list(deg)
    for i in range(len(deg)):
        deg[i] = deg[i] / 180 * 3.1415926
    deg = tuple(deg)
    return deg


def takeFirst(elem):
    return elem[0]


def Identify_TLC_Hight(picture_name,dx=200,dboundary=170,dinterval=100):
    image = imread(picture_name)
    #产生灰度图
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    x_shade=np.zeros([gray.shape[1]])
    x_white=[]
    for i in range(gray.shape[1]):
        x_shade[i] = np.sum(gray[:,i])/gray.shape[0]
        if x_shade[i]>40:
            x_white.append(i)
    #print(x_white[0],x_white[-1])

    x_1_center=x_white[0]+dboundary
    x_2_center=x_1_center+dx
    x_3_center=x_2_center+dx
    x_4_center=x_3_center+dx
    y_limit=1800
    y_1_shade=np.zeros(y_limit)
    y_2_shade=np.zeros(y_limit)
    y_3_shade=np.zeros(y_limit)
    y_4_shade=np.zeros(y_limit)
    x_1_shade=np.zeros([2*dinterval])
    x_2_shade=np.zeros([2*dinterval])
    x_3_shade=np.zeros([2*dinterval])
    x_4_shade=np.zeros([2*dinterval])
    for i in range(y_limit):
        y_1_shade[i]=np.sum(gray[i,x_white[0]:x_1_center+dinterval])/(2*dinterval)
        y_2_shade[i] = np.sum(gray[i, x_2_center - dinterval:x_2_center + dinterval]) / (2 * dinterval)
        y_3_shade[i] = np.sum(gray[i, x_3_center - dinterval:x_3_center + dinterval]) / (2 * dinterval)
        y_4_shade[i] = np.sum(gray[i, x_4_center - dinterval:x_white[-1]]) / (2 * dinterval)

    y_hight=np.array([np.where(y_1_shade==np.min(y_1_shade))[0][0],np.where(y_2_shade==np.min(y_2_shade))[0][0],
             np.where(y_3_shade==np.min(y_3_shade))[0][0],np.where(y_4_shade==np.min(y_4_shade))[0][0]])


    x_1_shade=gray[y_hight[0],x_1_center-dinterval:x_1_center+dinterval]
    x_2_shade = gray[y_hight[1], x_2_center - dinterval:x_2_center + dinterval]
    x_3_shade = gray[y_hight[2], x_3_center - dinterval:x_3_center + dinterval]
    x_4_shade = gray[y_hight[3], x_4_center - dinterval:x_white[-1]]
    x_hight_1=np.array([x_1_center-dinterval+np.where(x_1_shade==np.min(x_1_shade))[0][0],x_2_center-dinterval+np.where(x_2_shade==np.min(x_2_shade))[0][0],
             x_3_center-dinterval+np.where(x_3_shade==np.min(x_3_shade))[0][0],x_4_center-dinterval+np.where(x_4_shade==np.min(x_4_shade))[0][0]])

    for i in range(2 * dinterval):
        x_1_shade[i] = np.sum(gray[:, x_1_center - dinterval + i]) / (gray.shape[0])
        x_2_shade[i] = np.sum(gray[:, x_2_center - dinterval + i]) / (gray.shape[0])
        x_3_shade[i] = np.sum(gray[:, x_3_center - dinterval + i]) / (gray.shape[0])
    for i in range(dinterval):
        x_4_shade[i] = np.sum(gray[:, x_4_center - int(dinterval/2) + i]) / (gray.shape[0])

    x_hight_2 = np.array([x_1_center - dinterval + np.where(x_1_shade == np.min(x_1_shade))[0][0],
                          x_2_center - dinterval + np.where(x_2_shade == np.min(x_2_shade))[0][0],
                          x_3_center - dinterval + np.where(x_3_shade == np.min(x_3_shade))[0][0],
                          x_4_center - dinterval + np.where(x_4_shade == np.min(x_4_shade))[0][0]])
    x_hight_3=np.array([x_1_center,x_2_center,x_3_center,x_4_center])
    x_hight=(x_hight_1+x_hight_3)/2
    true_hight=(gray.shape[0]-y_hight)*5.30/196+1.38-8
    for i in range(true_hight.shape[0]):
        if true_hight[i]<0:
            true_hight[i]=0
    #画图
    # plt.subplot(111), plt.imshow(image)
    # plt.scatter(x_hight,y_hight,c='red')
    # plt.show()
    return list(true_hight)
def Identify_Eluent_Hight(picture_name):
    #print(Identify_TLC_true(r"C:\Users\xxhhss6910\PycharmProjects\AUBO\Save_Fig\Experiment_8171117\board_2.jpg"))
    image = imread(picture_name)
    #产生灰度图
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    y_shade = np.zeros([gray.shape[0]])
    for i in range(gray.shape[0]):
        y_shade[i] = np.sum(gray[i, :]) / (gray.shape[1])
    y_hight=np.where(y_shade==np.max(y_shade))[0][0]-55  #上边缘
    hight=(1944-y_hight)*0.0267+17.82-8
    return hight

    #685--1101 11.12mm  1pix=0.0267 mm
    #remain 17.82mm
def Calculate_RF(tlc_hight,eluent_hight):
    RF=[]
    for i in range(len(tlc_hight)):
        RF.append(tlc_hight[i]/eluent_hight)
    return RF


def write_xlsx(TLC_hight, Eluent_hight, RF):
    x1 = openpyxl.load_workbook('tb_tlc.xlsx')  # 找到需要xlsx文件的位置
    sheet_name = x1.get_sheet_names()
    sheet_info = x1.get_sheet_by_name(sheet_name[0])
    sheet_CAS = x1.get_sheet_by_name(sheet_name[1])
    CAS_rows = sheet_CAS.max_row
    CAS_NO_list = []
    for i in range(2, CAS_rows + 1):
        CAS_NO_list.append(sheet_CAS.cell(i, 1).value)
    print(CAS_NO_list)

    for j in range(2):
        sheet_info.cell(j + 2, 5).value = CAS_NO_list[4 * j]
        sheet_info.cell(j + 2, 6).value = CAS_NO_list[4 * j + 1]
        sheet_info.cell(j + 2, 7).value = CAS_NO_list[4 * j + 2]
        sheet_info.cell(j + 2, 8).value = CAS_NO_list[4 * j + 3]
        sheet_info.cell(j + 2, 9).value = Eluent_hight[j]
        sheet_info.cell(j + 2, 10).value = TLC_hight[j][0]
        sheet_info.cell(j + 2, 11).value = TLC_hight[j][1]
        sheet_info.cell(j + 2, 12).value = TLC_hight[j][2]
        sheet_info.cell(j + 2, 13).value = TLC_hight[j][3]
        sheet_info.cell(j + 2, 14).value = RF[j][0]
        sheet_info.cell(j + 2, 15).value = RF[j][1]
        sheet_info.cell(j + 2, 16).value = RF[j][2]
        sheet_info.cell(j + 2, 17).value = RF[j][3]

    x1.save('tb_tlc.xlsx')
    return 0
# ----------------AUBO SET----------------
robot, ret = robot_login()
# 速度和加速度设置，默认0.03
# robot.set_end_max_line_acc(0.01)
# robot.set_end_max_line_velc(0.03)
# robot.set_end_max_line_acc(0.10)
# robot.set_end_max_line_velc(0.21)

robot.set_end_max_line_acc(0.20)
robot.set_end_max_line_velc(0.42)
robot.set_end_max_line_acc(0.20)
robot.set_end_max_line_velc(0.42)
joint_acc=2.5
joint_velc=2.5
robot.set_joint_maxacc([joint_acc,joint_acc,joint_acc,joint_acc,joint_acc,joint_acc])
robot.set_joint_maxvelc([joint_velc,joint_velc,joint_velc,joint_velc,joint_velc,joint_velc])
print("小遨准备就绪！")

# -----------------DOBOT SET----------------
api = dobotSDK.load()
timer_running = True

resultConnect = dobotSDK.ConnectDobot(api, "192.168.1.6")
# print("resultConnect", resultConnect)

dobotSDK.SetControlMode(api, 1)
# 等待使能成功
# Wait Enable ControlMode Success
dobotSDK.dSleep(8000)
print("小越准备就绪！")

# ------------------夹爪SET---------------------
ser = serial.Serial('COM7', 115200)
ser.timeout = 0.01
ser.isOpen()
setid(7)
for i in range(1, 255):
    if getid(i) == 7:
        gripper_id = i
        break
print("夹爪准备就绪！")




#==================AUBO的一些位置=================
#-----------初始位置------------
Grab_initial_rad = deg_to_rad((0, -7.292, -75.694, 21.596, -90, 0))

#-------夹中间的板相关位置------------
# 规划点1
Grab_rad_1 = deg_to_rad((12.779, 2.291, -83.133, 91.007, 13.963, -84.814))
# #规划点2（左）
Grab_left_rad = deg_to_rad((14.331, 23.589, -119.055, 34.173, 15.513, -85.195))
# 夹板
Grab_rad = deg_to_rad((26.769, 29.753, -108.181, 40.687, 27.941, -85.868))
# 提上去
Grab_up_rad = deg_to_rad((26.769, 11.083, -78.062, 89.475, 27.941, -85.868))

#-------夹左边板相关位置-----------------
# 规划点1
Grab_rad_1_left = deg_to_rad((16.029, -1.389, -88.0854, 90.426, 17.208, -85.513))
# #规划点2（左）
Grab_left_rad_left = deg_to_rad((16.029, 21.054, -123.339, 32.729, 17.208, -85.513))
# 夹板
Grab_rad_left = deg_to_rad((28.624, 27.450, -112.301, 38.949, 29.796, -85.958))
# 提上去
Grab_up_rad_left = deg_to_rad((28.624, 7.750, -81.892, 89.058, 29.796, -85.958))


#-------夹右边板相关位置----------------
Grab_left_rad_right = deg_to_rad((20.5099, 29.622, -108.410, 39.663, 21.684, -86.121))
Grab_rad_right = deg_to_rad((25.061, 31.938, -103.555, 43.043, 26.234, -85.775))
Grab_up_rad_right = deg_to_rad((25.061, 14.724, -73.005, 90.808, 26.234, -85.775))
#----------吸盘相关设置-----------
# 最前面的盘
Blow_up_rad_1 = deg_to_rad((16.272, 13.491, -74.843, 90.300, 104.792, 0.517))
Blow_rad_1 = deg_to_rad((16.272, 19.633, -97.728, 61.272, 104.8, 0.517))

#第二个
Blow_up_rad_2 = deg_to_rad((7.2338, 8.584, -80.820, 89.268, 95.757, 0.733))
Blow_rad_2 = deg_to_rad((7.2338, 15.332, -103.6, 59.717, 95.757, 0.733))

#第三个
Blow_up_rad_3 = deg_to_rad((-2.983, 5.108, -84.746, 88.821, 85.542, 0.969))
Blow_rad_3 = deg_to_rad((-2.983, 12.188, -107.638, 58.848, 85.542, 0.969))

#第四个
Blow_up_rad_4 = deg_to_rad((-14.773, 3.419, -86.545, 88.660, 73.755, 1.251))
Blow_rad_4 = deg_to_rad((-14.773, 10.670, -109.584, 58.367, 73.755, 1.251))

#--------------放板相关设置-----------------
# 第一个瓶
# 规划点1
Put_high_rad_1 = deg_to_rad((13.267, 15.520, -73.258, 1.219, -90, 13.266))
# 瓶口位置（规划点2） z=0.333
Put_up_rad_1 = deg_to_rad((13.267, 16.223, -83.042, -9.269, -90, 13.266))
# 放板位置
Put_rad_1 = deg_to_rad((13.267, 18.532, -90.518, -19.054, -90, 13.266))
#调整位置(往右偏)
Put_adjust_1=deg_to_rad((12.886, 18.305, -90.829, -19.137, -90, 12.885))


#第二个瓶
Put_high_rad_2 = deg_to_rad((5.832, 11.417, -78.368, 0.2116, -90, 5.831))
Put_up_rad_2 = deg_to_rad((6.059, 12.406, -89.726, -12.136, -90, 6.058))
Put_rad_2 = deg_to_rad((6.059, 14.378, -95.343, -19.724, -90, 6.058))
Put_adjust_2=deg_to_rad((5.649, 14.160, -95.460, -19.624, -90, 5.648))

#第三个瓶
Put_high_rad_3 = deg_to_rad((-1.754, 8.749, -81.640, -0.391, -90, -1.755))
Put_up_rad_3 = deg_to_rad((-1.931, 9.989, -93.932, -13.923, -90, -1.932))
Put_rad_3 = deg_to_rad((-1.931, 12.033, -99.119, -21.154, -90, -1.932))
Put_adjust_3=deg_to_rad((-2.314, 11.943, -99.231, -21.176, -90, -2.315))

#第四个瓶
Put_high_rad_4 = deg_to_rad((-10.833, 7.343, -83.097, -0.444, -90, -10.834))
Put_up_rad_4 = deg_to_rad((-10.833, 8.341, -92.774, -11.119, -90, -10.834))
Put_rad_4 = deg_to_rad((-10.833, 13.068, -103.641, -26.713, -90, -10.834))
Put_adjust_4=deg_to_rad((-11.363, 13.047, -103.667, -26.719, -90, -11.363))
#补充板子相关设置
#规划点1（高）
Xi_high = deg_to_rad((42.866, 46.956, -35.680, 97.425, 80.367, 0.009247))
#吸新版子
Xi_low = deg_to_rad((42.866, 47.321, -51.907, 80.835, 80.367, 0.009247))
#放板子_1（右边板子）
Xi_fang_high_1=deg_to_rad((53.170, 32.047, -60.829, 87.186, 90.671, 0.02))
Xi_fang_1=deg_to_rad((53.170, 35.889, -74.323, 69.850, 90.671, 0.02))
#放板子_2（中间板子）
Xi_fang_high_2=deg_to_rad((50.730, 34.475, -57.045, 88.540, 88.231, 0.0170))
Xi_fang_2=deg_to_rad((50.678, 37.899, -70.995, 71.161, 88.178, 0.0155))
#放板子_3 (右边板子)
Xi_fang_high_3=deg_to_rad((48.4023, 37.4899, -45.9078, 96.6617, 85.903, 0.0143))
Xi_fang_3=deg_to_rad((48.4023, 40.126, -67.230, 72.701, 85.903, 0.0141))
#----------------第一个拍照位置--------------
Photo_line_rad = deg_to_rad((-32.100, 12.410, -72.029, 5.838, -90.037, -32.102))

#--------------第二个拍照位置------------
Photo_rad_prepare = deg_to_rad((118.487, 14.399, -133.905, 30.7047, 119.641, -88.769))
Lu_rad_1 = deg_to_rad((116.0390, 21.528, -135.642, 22.434, 116.587, -88.436))
Lu_rad_2 = deg_to_rad((108.269, 27.727, -119.142, 32.996, 108.823, -88.312))
Lu_rad_3=deg_to_rad((108.269, 24.244, -118.466, 37.156, 108.823, -88.312))
#----------------------丢弃位置----------------

Remove_rad_high=deg_to_rad((16.028, -1.450, -89.664, 88.907, 17.207, -85.512))
Remove_rad_up=deg_to_rad((83.295, 27.371, -43.712, 18.915, -90, 5.962))
Remove_rad=deg_to_rad((83.295, 24.699, -57.802, 7.496, -90, 5.962))

#------------------化合物编号-------------
Compound_index=np.array([[1,2,3,4],[5,6,7,8],
                         [9,10,11,12],[13,14,15,16],
                         [377,378,379,380],[381,382,383,384]])
Eluent_info=['PE','EA']
Eluent_value=np.array([[1,0],
                       [50,1],
                       [20,1]])

#创建文件夹
Dir_name=str(datetime.datetime.now().month)+str(datetime.datetime.now().day)+str(datetime.datetime.now().hour)+str(datetime.datetime.now().minute)
#save model dir
try:
    os.makedirs('Save_Fig/Experiment_%s'%(Dir_name))
except OSError:
    pass

#----------封装函数-----------------
def Photo_TLC(ret,group_num,num):
    if num==1:
        ret = robot.move_joint(Blow_up_rad_1)
        ret = robot.move_line(Blow_rad_1)
        robot.set_board_io_status(5, 'U_DO_01', 0)
        ret = robot.move_line(Blow_up_rad_1)

    if num==2:
        ret = robot.move_joint(Blow_up_rad_2)
        ret = robot.move_line(Blow_rad_2)
        robot.set_board_io_status(5, 'U_DO_01', 0)
        ret = robot.move_line(Blow_up_rad_2)

    if num==3:
        ret = robot.move_joint(Blow_up_rad_3)
        ret = robot.move_line(Blow_rad_3)
        robot.set_board_io_status(5, 'U_DO_01', 0)
        ret = robot.move_line(Blow_up_rad_3)

    if num==4:
        ret = robot.move_joint(Blow_up_rad_4)
        ret = robot.move_line(Blow_rad_4)
        robot.set_board_io_status(5, 'U_DO_01', 0)
        ret = robot.move_line(Blow_up_rad_4)


    # 送去拍照
    print("送到紫外灯下观察")
    ret = robot.move_joint(Grab_rad_1)
    ret = robot.move_line(Grab_left_rad)
    # 规划点2（注意容易撞）
    ret = robot.move_line(Photo_rad_prepare)
    # 伸进去拍照
    ret = robot.move_line(Lu_rad_1)
    ret = robot.move_line(Lu_rad_2)
    ret = robot.move_line(Lu_rad_3)
    print("打开紫外灯")
    robot.set_board_io_status(5, 'U_DO_02', 1)
    time.sleep(2)
    print("拍照记录")
    take_photo(file_name=f"Save_Fig/Experiment_{Dir_name}/{Compound_index[group_num-1][0]}-{Compound_index[group_num-1][1]}-{Compound_index[group_num-1][2]}-{Compound_index[group_num-1][3]}_{Eluent_info[0]}-{Eluent_value[num-1][0]}_{Eluent_info[1]}-{Eluent_value[num-1][1]}", camera_index=1)
    print("关闭紫外灯")
    robot.set_board_io_status(5, 'U_DO_02', 0)
    print("记录完毕")

    # board_path = 'Save_Fig/Experiment_%s/' % (Dir_name) + 'board_%d_%d.jpg'%(Dirnum)
    # line_path = 'Save_Fig/Experiment_%s/' % (Dir_name) + 'line_%d.jpg'%(num)
    # TLC_hight = Identify_TLC_Hight(board_path)
    # Eluent_hight = Identify_Eluent_Hight(line_path)
    # RF = Calculate_RF(TLC_hight, Eluent_hight)

    # 拍照
    ret = robot.move_line(Photo_rad_prepare)
    ret = robot.move_line(Grab_left_rad)
    ret = robot.move_line(Grab_rad_1)
    #丢垃圾
    ret = robot.move_line(Remove_rad_high)
    ret = robot.move_joint(Remove_rad_up)
    ret = robot.move_line(Remove_rad)
    movetgt(250)
    ret = robot.move_joint(Remove_rad_up)
    ret = robot.move_joint(Grab_initial_rad)
    print("流程结束")
    return 0

Dobot_initial_pos = [220.22, 6.8992, 120, 0, 0, 0]
resultStart = dobotSDK.MovJ(api, Dobot_initial_pos, isBlock=True)

ROW_NUM = 8
# 有几行
COL_NUM = 1
# 间距
Dy=-17.471
Dx=-17.580
# 补偿量
COMPENSATE_Y=-0.0513
COMPENSATE_X=0.0474
DIP_HIGHT=64
Dobot_first_tube_high = [84.155,-232.9, 120, 0, 0, 0]
Wash_high = [-69.95, -362.6, 120, 0, 0, 0]
Wash_low = [-69.95, -362.6, DIP_HIGHT,0, 0, 0]
birang_pos = [-3.237, -251.4, 135.49, 0, 0, 0]
COUNT=0    #板上点的点的count
COUNT_TLC=0  #点了几个板
TLC_NUM=4
NUM=0
WAITING_TIME=300 #s
time_record=[0,0,0,0]

print("开始点板")
Group_num=0
#col=1
for col in range(1,4):
    for col_iter in range(2):
        Group_num+=1
        if col_iter==0:
            row_index=[0,1,2,3]
        if col_iter==1:
            row_index=[4,5,6,7]
        COUNT = 0
        if Group_num==1:
            for row in row_index:
                Dobot_tube_high = Dobot_first_tube_high.copy()
                Dobot_tube_high[0] += COMPENSATE_Y * Dy * row + Dx * col
                Dobot_tube_high[1] += Dy * row + COMPENSATE_X * Dx * col
                Dobot_tube_low = Dobot_tube_high.copy()
                Dobot_tube_low[2] = DIP_HIGHT
                resultStart = dobotSDK.MovJ(api, Dobot_tube_high, isBlock=True)
                resultStart = dobotSDK.MovL(api, Dobot_tube_low, isBlock=True)
                resultStart = dobotSDK.MovL(api, Dobot_tube_high, isBlock=True)
                for iter in range(3):
                    # 前四个管子点在右边的板上,后面四个板子点在左边的板上
                    if iter == 0:
                        # 左边的板子
                        Dobot_TLC_initial = [249.44, 109.53, 6.4, 0, 0, 0]
                        Dobot_TLC_initial_high = [249.44, 109.53, 120, 0, 0, 0]
                        Dobot_TLC_initial_middle = [249.44, 109.53, 20, 0, 0, 0]

                    if iter == 1:
                        # 中间的板子
                        Dobot_TLC_initial = [248.42, 141.47, 6.4, 0, 0, 0]
                        Dobot_TLC_initial_high = [248.42, 141.47, 120, 0, 0, 0]
                        Dobot_TLC_initial_middle = [248.42, 141.47, 20, 0, 0, 0]

                    if iter == 2:
                        # 右边的板子
                        Dobot_TLC_initial = [246.37, 173, 6.9454, 0, 0, 0]
                        Dobot_TLC_initial_high = [246.37, 173, 120, 0, 0, 0]
                        Dobot_TLC_initial_middle = [246.37, 173, 20, 0, 0, 0]

                    TLC_Dy = 5.30
                    Dobot_TLC_high = Dobot_TLC_initial_high.copy()
                    Dobot_TLC_high[1] += (-COUNT * TLC_Dy)
                    Dobot_TLC_middle = Dobot_TLC_high.copy()
                    Dobot_TLC_middle[2] = 20
                    Dobot_TLC_low = Dobot_TLC_high.copy()
                    if iter == 0:
                        Dobot_TLC_low[2] =1.59
                    if iter == 1:
                        Dobot_TLC_low[2] =1.63
                    if iter == 2:
                        Dobot_TLC_low[2] = 2.0
                    resultStart = dobotSDK.MovJ(api, Dobot_TLC_high, isBlock=True)
                    resultStart = dobotSDK.MovJ(api, Dobot_TLC_middle, isBlock=True)
                    resultStart = dobotSDK.MovL(api, Dobot_TLC_low, isBlock=True)
                    if iter==2:
                        time.sleep(0.1)
                    resultStart = dobotSDK.MovL(api, Dobot_TLC_high, isBlock=True)

                # 吸水
                Dip_low_z =69
                Dip_high = [-89.87 + 10 * np.random.uniform(-1, 1), -305.4 + 30 * np.random.uniform(-1, 1), 120, 0, 0, 0]
                Dip_low = [Dip_high[0], Dip_high[1], Dip_low_z, 0, 0, 0]
                resultStart = dobotSDK.MovJ(api, Dip_high, isBlock=True)
                resultStart = dobotSDK.MovL(api, Dip_low, isBlock=True)
                time.sleep(2)
                resultStart = dobotSDK.MovL(api, Dip_high, isBlock=True)

                for k in range(2):
                    # 洗管
                    resultStart = dobotSDK.MovJ(api, Wash_high, isBlock=True)
                    resultStart = dobotSDK.MovL(api, Wash_low, isBlock=True)
                    resultStart = dobotSDK.MovL(api, Wash_high, isBlock=True)

                    # 吸水
                    Dip_high = [-89.87 + 10 * np.random.uniform(-1, 1), -305.4 + 30 * np.random.uniform(-1, 1), 120, 0, 0, 0]
                    Dip_low = [Dip_high[0], Dip_high[1], Dip_low_z, 0, 0, 0]
                    resultStart = dobotSDK.MovJ(api, Dip_high, isBlock=True)
                    resultStart = dobotSDK.MovL(api, Dip_low, isBlock=True)
                    time.sleep(2)
                    resultStart = dobotSDK.MovL(api, Dip_high, isBlock=True)

                COUNT+=1
        print("点板完成")
        resultStart = dobotSDK.MovJ(api, birang_pos, isBlock=True)
        for iter in range(3):

            # 大机器人开始行动
            # 先是夹左边的板子
            print("夹取TLC板放到展缸中")
            if iter == 0:
                movetgt(500)
                # 夹板（左)
                ret = robot.move_joint(Grab_rad_1_left)
                ret = robot.move_line(Grab_left_rad_left)
                ret = robot.move_line(Grab_rad_left)
                movetgt(0)
                time.sleep(2)
                ret = robot.move_line(Grab_up_rad_left)

            if iter == 1:
                # 夹板(中)
                movetgt(500)
                # 初始位置
                ret = robot.move_joint(Grab_rad_1)
                ret = robot.move_line(Grab_left_rad)
                ret = robot.move_line(Grab_rad)
                movetgt(0)
                time.sleep(2)
                ret = robot.move_line(Grab_up_rad)

            if iter == 2:
                # 夹板(右)
                movetgt(500)
                # 初始位置
                ret = robot.move_joint(Grab_rad_1)
                ret = robot.move_line(Grab_left_rad)
                ret = robot.move_line(Grab_left_rad_right)
                ret = robot.move_line(Grab_rad_right)
                movetgt(0)
                time.sleep(2)
                ret = robot.move_line(Grab_up_rad_right)

            print("板已夹起")
            print("打开瓶盖")
            if iter == 0:
                ret = robot.move_joint(Blow_up_rad_1)
                ret = robot.move_line(Blow_rad_1)
                robot.set_board_io_status(5, 'U_DO_01', 1)
                ret = robot.move_line(Blow_up_rad_1)
                print("将板放入展缸")
                ret = robot.move_joint(Put_high_rad_1)
                ret = robot.move_line(Put_up_rad_1)
                ret = robot.move_line(Put_rad_1)
                movetgt(100)
                ret = robot.move_line(Put_adjust_1)
                time_1 = time.time()
                ret = robot.move_line(Put_up_rad_1)
                ret = robot.move_line(Put_high_rad_1)
                print("盖上瓶盖")
                ret = robot.move_joint(Blow_up_rad_1)
                ret = robot.move_line(Blow_rad_1)
                robot.set_board_io_status(5, 'U_DO_01', 0)
                ret = robot.move_line(Blow_up_rad_1)
                time.sleep(27)
            if iter == 1:
                ret = robot.move_joint(Blow_up_rad_2)
                ret = robot.move_line(Blow_rad_2)
                robot.set_board_io_status(5, 'U_DO_01', 1)
                ret = robot.move_line(Blow_up_rad_2)
                print("将板放入展缸")
                ret = robot.move_joint(Put_high_rad_2)
                ret = robot.move_line(Put_up_rad_2)
                ret = robot.move_line(Put_rad_2)
                movetgt(100)
                ret = robot.move_line(Put_adjust_2)
                time_2 = time.time()
                ret = robot.move_line(Put_up_rad_2)
                ret = robot.move_line(Put_high_rad_2)
                print("盖上瓶盖")
                ret = robot.move_joint(Blow_up_rad_2)
                ret = robot.move_line(Blow_rad_2)
                robot.set_board_io_status(5, 'U_DO_01', 0)
                ret = robot.move_line(Blow_up_rad_2)
                time.sleep(27)
            if iter == 2:
                ret = robot.move_joint(Blow_up_rad_3)
                ret = robot.move_line(Blow_rad_3)
                robot.set_board_io_status(5, 'U_DO_01', 1)
                ret = robot.move_line(Blow_up_rad_3)
                print("将板放入展缸")
                ret = robot.move_joint(Put_high_rad_3)
                ret = robot.move_line(Put_up_rad_3)
                ret = robot.move_line(Put_rad_3)
                movetgt(100)
                ret = robot.move_line(Put_adjust_3)
                time_3 = time.time()
                ret = robot.move_line(Put_up_rad_3)
                ret = robot.move_line(Put_high_rad_3)
                print("盖上瓶盖")
                ret = robot.move_joint(Blow_up_rad_3)
                ret = robot.move_line(Blow_rad_3)
                robot.set_board_io_status(5, 'U_DO_01', 0)
                ret = robot.move_line(Blow_up_rad_3)

            ret = robot.move_joint(Grab_initial_rad)
            print("开始展开，请等待...")




        # 放板子
        ret = robot.move_joint(Xi_high)
        ret = robot.move_line(Xi_low)
        robot.set_board_io_status(5, 'U_DO_01', 1)
        ret = robot.move_line(Xi_high)
        # 放第一块板（左）
        ret = robot.move_line(Xi_fang_high_1)
        ret = robot.move_line(Xi_fang_1)
        robot.set_board_io_status(5, 'U_DO_01', 0)
        time.sleep(0.5)
        ret = robot.move_line(Xi_fang_high_1)
        # 吸板子
        ret = robot.move_line(Xi_high)
        ret = robot.move_line(Xi_low)
        robot.set_board_io_status(5, 'U_DO_01', 1)
        ret = robot.move_line(Xi_high)
        # 放第二块板
        ret = robot.move_line(Xi_fang_high_2)
        ret = robot.move_line(Xi_fang_2)
        robot.set_board_io_status(5, 'U_DO_01', 0)
        time.sleep(0.5)
        ret = robot.move_line(Xi_fang_high_2)

        # 吸板子
        ret = robot.move_line(Xi_high)
        ret = robot.move_line(Xi_low)
        robot.set_board_io_status(5, 'U_DO_01', 1)
        ret = robot.move_line(Xi_high)

        # 放第三块板
        ret = robot.move_line(Xi_fang_high_3)
        ret = robot.move_line(Xi_fang_3)
        robot.set_board_io_status(5, 'U_DO_01', 0)
        time.sleep(0.5)
        ret = robot.move_line(Xi_fang_high_3)
        # 回到初始位置
        ret = robot.move_joint(Grab_initial_rad)
        COUNT = 0
        if Group_num < 6:
            if col_iter==0:
                new_col=col
                row_index=[4,5,6,7]
            if col_iter==1:
                new_col=col+1
                row_index=[0,1,2,3]
            for row in row_index:
                Dobot_tube_high = Dobot_first_tube_high.copy()
                Dobot_tube_high[0] += COMPENSATE_Y * Dy * row + Dx * new_col
                Dobot_tube_high[1] += Dy * row + COMPENSATE_X * Dx * new_col
                Dobot_tube_low = Dobot_tube_high.copy()
                Dobot_tube_low[2] = DIP_HIGHT
                resultStart = dobotSDK.MovJ(api, Dobot_tube_high, isBlock=True)
                resultStart = dobotSDK.MovL(api, Dobot_tube_low, isBlock=True)
                resultStart = dobotSDK.MovL(api, Dobot_tube_high, isBlock=True)
                for iter in range(3):
                    # 前四个管子点在右边的板上,后面四个板子点在左边的板上
                    if iter == 0:
                        # 左边的板子
                        Dobot_TLC_initial = [249.44, 109.53, 6.4, 0, 0, 0]
                        Dobot_TLC_initial_high = [249.44, 109.53, 120, 0, 0, 0]
                        Dobot_TLC_initial_middle = [249.44, 109.53, 20, 0, 0, 0]

                    if iter == 1:
                        # 中间的板子
                        Dobot_TLC_initial = [248.42, 141.47, 6.4, 0, 0, 0]
                        Dobot_TLC_initial_high = [248.42, 141.47, 120, 0, 0, 0]
                        Dobot_TLC_initial_middle = [248.42, 141.47, 20, 0, 0, 0]

                    if iter == 2:
                        # 右边的板子
                        Dobot_TLC_initial = [246.37, 173, 6.9454, 0, 0, 0]
                        Dobot_TLC_initial_high = [246.37, 173, 120, 0, 0, 0]
                        Dobot_TLC_initial_middle = [246.37, 173, 20, 0, 0, 0]

                    TLC_Dy = 5.30
                    Dobot_TLC_high = Dobot_TLC_initial_high.copy()
                    Dobot_TLC_high[1] += (-COUNT * TLC_Dy)
                    Dobot_TLC_middle = Dobot_TLC_high.copy()
                    Dobot_TLC_middle[2] = 20
                    Dobot_TLC_low = Dobot_TLC_high.copy()
                    if iter == 0:
                        Dobot_TLC_low[2] =1.59
                    if iter == 1:
                        Dobot_TLC_low[2] =1.63
                    if iter == 2:
                        Dobot_TLC_low[2] = 2.0
                    resultStart = dobotSDK.MovJ(api, Dobot_TLC_high, isBlock=True)
                    resultStart = dobotSDK.MovJ(api, Dobot_TLC_middle, isBlock=True)
                    resultStart = dobotSDK.MovL(api, Dobot_TLC_low, isBlock=True)
                    if iter==2:
                        time.sleep(0.1)
                    resultStart = dobotSDK.MovL(api, Dobot_TLC_high, isBlock=True)

                # 吸水
                Dip_low_z = 69
                Dip_high = [-89.87 + 10 * np.random.uniform(-1, 1), -305.4 + 30 * np.random.uniform(-1, 1), 120, 0,
                            0, 0]
                Dip_low = [Dip_high[0], Dip_high[1], Dip_low_z, 0, 0, 0]
                resultStart = dobotSDK.MovJ(api, Dip_high, isBlock=True)
                resultStart = dobotSDK.MovL(api, Dip_low, isBlock=True)
                time.sleep(2)
                resultStart = dobotSDK.MovL(api, Dip_high, isBlock=True)

                for k in range(2):
                    # 洗管
                    resultStart = dobotSDK.MovJ(api, Wash_high, isBlock=True)
                    resultStart = dobotSDK.MovL(api, Wash_low, isBlock=True)
                    resultStart = dobotSDK.MovL(api, Wash_high, isBlock=True)

                    # 吸水
                    Dip_high = [-89.87 + 10 * np.random.uniform(-1, 1), -305.4 + 30 * np.random.uniform(-1, 1), 120,
                                0, 0, 0]
                    Dip_low = [Dip_high[0], Dip_high[1], Dip_low_z, 0, 0, 0]
                    resultStart = dobotSDK.MovJ(api, Dip_high, isBlock=True)
                    resultStart = dobotSDK.MovL(api, Dip_low, isBlock=True)
                    time.sleep(2)
                    resultStart = dobotSDK.MovL(api, Dip_high, isBlock=True)

                COUNT += 1
            resultStart = dobotSDK.MovJ(api, birang_pos, isBlock=True)
            #开始拿板
        FLAG=1
        while FLAG==1:
            time_end=time.time()
            print(time_end-time_1,time_end-time_2,time_end-time_3)
            if time_end-time_1>WAITING_TIME:
                time_record[0]=time_end-time_1
                #取第一个板
                movetgt(250)
                ret = robot.move_joint(Blow_up_rad_1)
                ret = robot.move_line(Blow_rad_1)
                robot.set_board_io_status(5, 'U_DO_01', 1)
                ret = robot.move_line(Blow_up_rad_1)
                ret = robot.move_joint(Put_high_rad_1)
                ret = robot.move_line(Put_up_rad_1)
                ret = robot.move_line(Put_rad_1)
                movetgt(0)
                time.sleep(2)
                ret = robot.move_line(Put_up_rad_1)
                ret = robot.move_joint(Put_high_rad_1)
                Photo_TLC(ret,group_num=Group_num,num=1)

                #结束
                time_1=time_end+99999.0


            if time_end - time_2 > WAITING_TIME:
                time_record[1] = time_end - time_2
                # 取第二个板
                movetgt(250)
                ret = robot.move_joint(Blow_up_rad_2)
                ret = robot.move_line(Blow_rad_2)
                robot.set_board_io_status(5, 'U_DO_01', 1)
                ret = robot.move_line(Blow_up_rad_2)
                ret = robot.move_joint(Put_high_rad_2)
                ret = robot.move_line(Put_up_rad_2)
                ret = robot.move_line(Put_rad_2)
                movetgt(0)
                time.sleep(2)
                ret = robot.move_line(Put_up_rad_2)
                ret = robot.move_joint(Put_high_rad_2)
                Photo_TLC(ret,group_num=Group_num, num=2)


                time_2 = time_end + 99999.0

            if time_end-time_3>WAITING_TIME:
                time_record[0]=time_end-time_3
                #取第一个板
                movetgt(300)
                ret = robot.move_joint(Blow_up_rad_3)
                ret = robot.move_line(Blow_rad_3)
                robot.set_board_io_status(5, 'U_DO_01', 1)
                ret = robot.move_line(Blow_up_rad_3)
                ret = robot.move_joint(Put_high_rad_3)
                ret = robot.move_line(Put_up_rad_3)
                ret = robot.move_line(Put_rad_3)
                movetgt(0)
                time.sleep(2)
                ret = robot.move_line(Put_up_rad_3)
                ret = robot.move_joint(Put_high_rad_3)
                Photo_TLC(ret,group_num=Group_num,num=3)

                #结束
                time_3=time_end+99999.0
                FLAG = 0
